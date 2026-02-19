from pathlib import Path

import openpyxl
from loguru import logger

from app import SUPPORTED_LANGUAGES
from app.translation import TranslationSource, TranslationRow


# ── XLSX backend ──────────────────────────────────────────────────────────────
class XlsxTranslationSource(TranslationSource):
    def __init__(self, input_path: Path, output_path: Path, lang_codes: list[str]) -> None:
        self.input_path = input_path
        self.output_path = output_path
        self.lang_codes = lang_codes
        self._wb = None
        self._sheet = None
        self._col_map: dict[str, int] = {}  # lang_code → column index

    def describe(self) -> str:
        return f"Excel  {self.input_path} → {self.output_path}"

    def load(self) -> list[TranslationRow]:
        logger.info(f"Loading workbook: {self.input_path}")
        self._wb = openpyxl.load_workbook(self.input_path)
        self._sheet = self._wb.active

        # Map header values to column indices
        header: dict[str, int] = {
            self._sheet.cell(row=1, column=c).value: c
            for c in range(1, self._sheet.max_column + 1)
        }
        logger.debug(f"Sheet headers: {header}")

        # Resolve a column index for each requested language
        for code in self.lang_codes:
            name, _ = SUPPORTED_LANGUAGES[code]
            col = header.get(code) or header.get(name)
            if col is None:
                logger.warning(f"[{code}] No column found for '{code}' or '{name}' in sheet.")
            else:
                self._col_map[code] = col

        rows: list[TranslationRow] = []
        for row in self._sheet.iter_rows(min_row=2):
            key = row[0].value
            source_text = row[1].value
            translations = {
                code: (self._sheet.cell(row=row[0].row, column=col).value)
                for code, col in self._col_map.items()
            }
            rows.append(TranslationRow(key=key, source_text=source_text, translations=translations))

        logger.info(f"Loaded {len(rows)} rows from sheet.")
        return rows

    def save(self, rows: list[TranslationRow]) -> None:
        for sheet_row in self._sheet.iter_rows(min_row=2):
            key = sheet_row[0].value
            match = next((r for r in rows if r.key == key), None)
            if match is None:
                continue
            for code, col in self._col_map.items():
                value = match.translations.get(code)
                if value is not None:
                    self._sheet.cell(row=sheet_row[0].row, column=col).value = value

        self._wb.save(self.output_path)
        logger.success(f"Workbook saved → {self.output_path}")

